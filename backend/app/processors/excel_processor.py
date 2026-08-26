"""
ExcelProcessor — FileProcessor implementation for XLSX/XLS workbooks.

Produces BOTH structured and unstructured regions from a single workbook.
Each sheet is inspected independently and may yield regions with different
classifications (STRUCTURED, SEMI_STRUCTURED, UNSTRUCTURED, or IGNORE).
"""

from __future__ import annotations

import asyncio
import logging
from datetime import date, datetime, time
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.errors.ingestion_errors import FileProcessingError
from app.processors.protocol import (
    ColumnSchema,
    DetectedRegion,
    HeaderDetectionResult,
    InspectionResult,
    NormalizedDataset,
    SheetInfo,
    ValidationWarning,
)

logger = logging.getLogger(__name__)

# Configuration constants
_MAX_SAMPLE_ROWS = 20
_MIN_REGION_ROWS = 1
_BLANK_ROW_THRESHOLD = 2  # Consecutive blank rows separate regions
_BLANK_COL_THRESHOLD = 2  # Consecutive blank columns separate regions


class ExcelProcessor:
    """FileProcessor implementation for XLSX/XLS workbooks.

    Produces BOTH structured and unstructured regions from a single workbook.
    Each sheet is inspected independently and may yield regions with different
    classifications.
    """

    def can_process(self, file_type: str) -> bool:
        """Return True if this processor handles the given file type.

        Args:
            file_type: File extension identifier.

        Returns:
            True for "xlsx" and "xls" file types.
        """
        return file_type in ("xlsx", "xls")

    async def inspect(self, file_path: str) -> InspectionResult:
        """Discover sheets, detect regions, and provide content samples for classification.

        Offloads all synchronous openpyxl I/O to a thread pool via
        asyncio.to_thread so the FastAPI event loop is never blocked.
        Large workbooks (thousands of rows) can be inspected without
        causing HTTP request timeouts.

        Args:
            file_path: Path to the XLSX/XLS file on disk.

        Returns:
            InspectionResult with detected regions and workbook metadata.

        Raises:
            FileProcessingError: If the workbook cannot be opened or inspected.
        """
        return await asyncio.to_thread(self._sync_inspect, file_path)

    def _sync_inspect(self, file_path: str) -> InspectionResult:
        """Synchronous implementation of inspect() — runs inside asyncio.to_thread."""
        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as exc:
            raise FileProcessingError(
                file_name=file_path,
                message=f"Failed to open workbook: {exc}",
                detail=str(exc),
            ) from exc

        try:
            sheets = self._discover_sheets(workbook)
            all_regions: list[DetectedRegion] = []

            for sheet_info in sheets:
                sheet = workbook[sheet_info.name]
                regions = self._detect_regions(sheet)
                all_regions.extend(regions)

            # Build file-level metadata
            metadata: dict[str, str] = {
                "sheet_count": str(len(sheets)),
                "sheet_names": ", ".join(s.name for s in sheets),
            }
            if workbook.properties:
                if workbook.properties.creator:
                    metadata["author"] = workbook.properties.creator
                if workbook.properties.created:
                    metadata["created"] = str(workbook.properties.created)

            # Extract file name from path
            file_name = file_path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]

            return InspectionResult(
                file_name=file_name,
                file_type="xlsx",
                regions=all_regions,
                metadata=metadata,
            )
        except FileProcessingError:
            raise
        except Exception as exc:
            raise FileProcessingError(
                file_name=file_path,
                message=f"Error during workbook inspection: {exc}",
                detail=str(exc),
            ) from exc
        finally:
            workbook.close()

    async def extract(
        self, file_path: str, region: DetectedRegion | None = None
    ) -> NormalizedDataset:
        """Extract structured records from a confirmed region (non-blocking).

        Offloads all synchronous openpyxl I/O to asyncio.to_thread.

        Args:
            file_path: Path to the XLSX/XLS file on disk.
            region: The specific region to extract. If None, extracts the first
                    detected region from the first sheet.

        Returns:
            NormalizedDataset with columns, records, and source traceability.

        Raises:
            FileProcessingError: If extraction fails.
        """
        return await asyncio.to_thread(self._sync_extract, file_path, region)

    def _sync_extract(
        self, file_path: str, region: DetectedRegion | None = None
    ) -> NormalizedDataset:
        """Synchronous implementation of extract() — runs inside asyncio.to_thread."""
        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as exc:
            raise FileProcessingError(
                file_name=file_path,
                message=f"Failed to open workbook for extraction: {exc}",
                detail=str(exc),
            ) from exc

        try:
            if region is None:
                # Detect the first region from the workbook
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    regions = self._detect_regions(sheet)
                    if regions:
                        region = regions[0]
                        break
                if region is None:
                    raise FileProcessingError(
                        file_name=file_path,
                        message="No data regions detected in workbook",
                    )

            sheet = workbook[region.sheet_name]

            # Determine header row and infer schema
            header_result = self._detect_header(sheet, region)
            header_row = header_result.header_row
            columns = self._infer_schema(sheet, region, header_row)

            # Get column names from the header row
            column_names: list[str] = [col.name for col in columns]

            # Read data rows below the header
            data_start_row = header_row + 1
            records: list[dict[str, Any]] = []
            row_index = 0

            for row_idx in range(data_start_row, region.end_row):
                row_data: dict[str, Any] = {}
                for col_offset, col_name in enumerate(column_names):
                    col_idx = region.start_column + col_offset
                    cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                    value = self._get_cell_value(cell)
                    row_data[col_name] = value

                records.append(row_data)
                row_index += 1

            source_location = (
                f"{region.sheet_name}!R{region.start_row + 1}:R{region.end_row}"
                f"C{region.start_column + 1}:C{region.end_column}"
            )

            return NormalizedDataset(
                dataset_id=None,
                source_file_id="",
                sheet_name=region.sheet_name,
                columns=columns,
                records=records,
                classification="STRUCTURED",
                source_location=source_location,
                confidence=header_result.confidence,
                warnings=header_result.warnings,
            )
        except FileProcessingError:
            raise
        except Exception as exc:
            raise FileProcessingError(
                file_name=file_path,
                message=f"Error during data extraction: {exc}",
                detail=str(exc),
            ) from exc
        finally:
            workbook.close()

    async def extract_text(self, file_path: str, region: DetectedRegion) -> str:
        """Extract plain text from an unstructured region for RAG processing (non-blocking).

        Offloads all synchronous openpyxl I/O to asyncio.to_thread.

        Args:
            file_path: Path to the XLSX/XLS file on disk.
            region: The specific region to extract text from.

        Returns:
            Non-empty string for regions with content, empty string otherwise.

        Raises:
            FileProcessingError: If the workbook cannot be opened.
        """
        return await asyncio.to_thread(self._sync_extract_text, file_path, region)

    def _sync_extract_text(self, file_path: str, region: DetectedRegion) -> str:
        """Synchronous implementation of extract_text() — runs inside asyncio.to_thread."""
        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as exc:
            raise FileProcessingError(
                file_name=file_path,
                message=f"Failed to open workbook for text extraction: {exc}",
                detail=str(exc),
            ) from exc

        try:
            sheet = workbook[region.sheet_name]
            lines: list[str] = []

            for row_idx in range(region.start_row, region.end_row):
                row_parts: list[str] = []
                for col_idx in range(region.start_column, region.end_column):
                    cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                    value = self._get_cell_value(cell)
                    if value is not None:
                        text = str(value).strip()
                        if text:
                            row_parts.append(text)

                if row_parts:
                    lines.append(" ".join(row_parts))

            return "\n".join(lines)
        except Exception as exc:
            raise FileProcessingError(
                file_name=file_path,
                message=f"Error during text extraction: {exc}",
                detail=str(exc),
            ) from exc
        finally:
            workbook.close()

    def validate(self, normalized: NormalizedDataset) -> list[ValidationWarning]:
        """Validate a normalized dataset for consistency issues.

        Checks:
        - Column count consistency across records
        - Empty column names
        - Missing required values

        Args:
            normalized: The normalized dataset to validate.

        Returns:
            List of validation warnings (empty if no issues found).
        """
        warnings: list[ValidationWarning] = []

        # Check for empty column names
        for col in normalized.columns:
            if not col.name or not col.name.strip():
                warnings.append(
                    ValidationWarning(
                        field=f"column_{col.column_index}",
                        message=f"Column at index {col.column_index} has empty name",
                        severity="warning",
                    )
                )

        # Check for duplicate column names
        col_names = [c.name for c in normalized.columns]
        seen: set[str] = set()
        for name in col_names:
            if name in seen:
                warnings.append(
                    ValidationWarning(
                        field=name,
                        message=f"Duplicate column name: '{name}'",
                        severity="warning",
                    )
                )
            seen.add(name)

        # Check record consistency
        expected_cols = len(normalized.columns)
        for idx, record in enumerate(normalized.records):
            if len(record) != expected_cols:
                warnings.append(
                    ValidationWarning(
                        field=f"row_{idx}",
                        message=(
                            f"Row {idx} has {len(record)} columns, "
                            f"expected {expected_cols}"
                        ),
                        severity="warning",
                    )
                )
                # Only report first few mismatches
                if idx >= 5:
                    break

        return warnings

    # -------------------------------------------------------------------------
    # Internal methods
    # -------------------------------------------------------------------------

    def _discover_sheets(self, workbook: Any) -> list[SheetInfo]:
        """Get all sheet metadata from the workbook.

        Args:
            workbook: An openpyxl Workbook instance.

        Returns:
            List of SheetInfo for each sheet in the workbook.
        """
        sheets: list[SheetInfo] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Determine visibility
            visible = sheet.sheet_state == "visible"

            # Get dimensions
            row_count = sheet.max_row or 0
            column_count = sheet.max_column or 0

            # Count merged cells
            merged_cell_count = len(sheet.merged_cells.ranges) if hasattr(sheet, "merged_cells") and sheet.merged_cells else 0

            sheets.append(
                SheetInfo(
                    name=sheet_name,
                    visible=visible,
                    row_count=row_count,
                    column_count=column_count,
                    merged_cell_count=merged_cell_count,
                )
            )

        return sheets

    def _detect_regions(self, sheet: Any) -> list[DetectedRegion]:
        """Find contiguous populated rectangular areas within a sheet.

        Does NOT assume row 1 = header or one sheet = one dataset.
        Identifies separate regions by detecting blank row/column gaps.

        Args:
            sheet: An openpyxl Worksheet instance.

        Returns:
            List of DetectedRegion with content samples and raw_text.
        """
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        if max_row == 0 or max_col == 0:
            return []

        # Build a boolean grid of non-empty cells
        # Use 1-based indexing internally to match openpyxl
        non_empty_rows: list[int] = []

        for row_idx in range(1, max_row + 1):
            has_content = False
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = self._get_cell_value(cell)
                if value is not None and str(value).strip():
                    has_content = True
                    break
            if has_content:
                non_empty_rows.append(row_idx)

        if not non_empty_rows:
            return []

        # Split rows into contiguous groups separated by blank row gaps
        row_groups: list[list[int]] = []
        current_group: list[int] = [non_empty_rows[0]]

        for i in range(1, len(non_empty_rows)):
            gap = non_empty_rows[i] - non_empty_rows[i - 1] - 1
            if gap >= _BLANK_ROW_THRESHOLD:
                row_groups.append(current_group)
                current_group = [non_empty_rows[i]]
            else:
                current_group.append(non_empty_rows[i])

        if current_group:
            row_groups.append(current_group)

        # For each row group, determine column boundaries
        regions: list[DetectedRegion] = []
        sheet_name = sheet.title

        for group in row_groups:
            start_row_1based = group[0]
            end_row_1based = group[-1]

            # Find populated column range across all rows in the group
            min_col = max_col + 1
            max_col_found = 0

            for row_idx in group:
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = self._get_cell_value(cell)
                    if value is not None and str(value).strip():
                        min_col = min(min_col, col_idx)
                        max_col_found = max(max_col_found, col_idx)

            if max_col_found == 0:
                continue

            # Convert to 0-based indices for DetectedRegion
            start_row = start_row_1based - 1
            end_row = end_row_1based  # end_row is exclusive in our protocol
            start_column = min_col - 1
            end_column = max_col_found  # end_column is exclusive

            row_count = end_row - start_row
            column_count = end_column - start_column

            # Skip trivially small regions
            if row_count < _MIN_REGION_ROWS:
                continue

            # Further split by blank column gaps (detect multiple tables side-by-side)
            col_ranges = self._detect_column_ranges(
                sheet, group, min_col, max_col_found
            )

            for col_start, col_end in col_ranges:
                actual_start_col = col_start - 1  # 0-based
                actual_end_col = col_end  # exclusive, 0-based
                actual_col_count = actual_end_col - actual_start_col

                # Extract content sample (up to _MAX_SAMPLE_ROWS rows)
                content_sample = self._extract_content_sample(
                    sheet,
                    start_row_1based,
                    end_row_1based,
                    col_start,
                    col_end,
                )

                # Build raw_text for narrative detection
                raw_text = self._build_raw_text(
                    sheet,
                    start_row_1based,
                    end_row_1based,
                    col_start,
                    col_end,
                )

                # Detect header for this region
                region_for_header = DetectedRegion(
                    region_id="",
                    sheet_name=sheet_name,
                    start_row=start_row,
                    end_row=end_row,
                    start_column=actual_start_col,
                    end_column=actual_end_col,
                    header_row=None,
                    content_sample=content_sample,
                    row_count=row_count,
                    column_count=actual_col_count,
                    raw_text=raw_text,
                )

                header_result = self._detect_header(sheet, region_for_header)

                region_id = (
                    f"{sheet_name}_{start_row}_{end_row}"
                    f"_{actual_start_col}_{actual_end_col}"
                )

                regions.append(
                    DetectedRegion(
                        region_id=region_id,
                        sheet_name=sheet_name,
                        start_row=start_row,
                        end_row=end_row,
                        start_column=actual_start_col,
                        end_column=actual_end_col,
                        header_row=header_result.header_row if header_result.confidence > 0.3 else None,
                        content_sample=content_sample,
                        row_count=row_count,
                        column_count=actual_col_count,
                        raw_text=raw_text,
                    )
                )

        return regions

    def _detect_column_ranges(
        self,
        sheet: Any,
        row_group: list[int],
        min_col: int,
        max_col: int,
    ) -> list[tuple[int, int]]:
        """Detect separate column ranges (tables side-by-side) within a row group.

        Splits columns by blank column gaps of _BLANK_COL_THRESHOLD or more.

        Args:
            sheet: The worksheet.
            row_group: 1-based row indices of the group.
            min_col: Minimum 1-based column with content.
            max_col: Maximum 1-based column with content.

        Returns:
            List of (start_col, end_col) tuples (1-based, end is inclusive).
        """
        # Determine which columns have any content across the row group
        populated_cols: list[int] = []
        for col_idx in range(min_col, max_col + 1):
            has_content = False
            for row_idx in row_group:
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = self._get_cell_value(cell)
                if value is not None and str(value).strip():
                    has_content = True
                    break
            if has_content:
                populated_cols.append(col_idx)

        if not populated_cols:
            return []

        # Split populated columns by gaps
        ranges: list[tuple[int, int]] = []
        range_start = populated_cols[0]
        prev_col = populated_cols[0]

        for i in range(1, len(populated_cols)):
            gap = populated_cols[i] - prev_col - 1
            if gap >= _BLANK_COL_THRESHOLD:
                ranges.append((range_start, prev_col))
                range_start = populated_cols[i]
            prev_col = populated_cols[i]

        ranges.append((range_start, prev_col))

        return ranges

    def _detect_header(
        self, sheet: Any, region: DetectedRegion
    ) -> HeaderDetectionResult:
        """Identify the header row within a region using deterministic heuristics.

        Heuristics applied:
        - Text density: candidate row should be mostly text (not numeric)
        - Value uniqueness: header cells should be unique
        - Consistent data types below: rows following the header should have
          consistent types per column
        - Contiguous records: many rows below candidate with content

        Args:
            sheet: The worksheet.
            region: The region to analyze for header.

        Returns:
            HeaderDetectionResult with detected header row, confidence, and reason.
        """
        if region.row_count == 0:
            return HeaderDetectionResult(
                header_row=region.start_row,
                confidence=0.0,
                detection_reason="Empty region, no header detected",
                warnings=["Region has no rows"],
            )

        best_row = region.start_row
        best_score = 0.0
        best_reason = "Default to first row"
        warnings: list[str] = []

        # Check the first few rows as header candidates
        candidate_limit = min(5, region.row_count)

        for offset in range(candidate_limit):
            candidate_row = region.start_row + offset
            row_1based = candidate_row + 1

            # Get candidate row values
            candidate_values: list[str] = []
            for col_offset in range(region.column_count):
                col_1based = region.start_column + col_offset + 1
                cell = sheet.cell(row=row_1based, column=col_1based)
                value = self._get_cell_value(cell)
                candidate_values.append(str(value) if value is not None else "")

            score = self._score_header_candidate(
                sheet, region, candidate_row, candidate_values
            )

            if score > best_score:
                best_score = score
                best_row = candidate_row
                best_reason = self._explain_header_selection(
                    candidate_values, score, offset
                )

        # Assess confidence
        confidence = min(1.0, best_score)
        if confidence < 0.5:
            warnings.append(
                "Low confidence header detection — region may not have a clear header"
            )

        return HeaderDetectionResult(
            header_row=best_row,
            confidence=round(confidence, 4),
            detection_reason=best_reason,
            warnings=warnings,
        )

    def _score_header_candidate(
        self,
        sheet: Any,
        region: DetectedRegion,
        candidate_row: int,
        candidate_values: list[str],
    ) -> float:
        """Score a candidate row's likelihood of being a header.

        Args:
            sheet: The worksheet.
            region: The region context.
            candidate_row: 0-based row index of the candidate.
            candidate_values: String values of cells in the candidate row.

        Returns:
            Score between 0.0 and 1.0.
        """
        score = 0.0

        # Filter out empty cells
        non_empty = [v for v in candidate_values if v.strip()]
        if not non_empty:
            return 0.0

        # Signal 1: Text density (non-numeric cells)
        text_count = sum(1 for v in non_empty if not self._looks_numeric(v))
        text_ratio = text_count / len(non_empty) if non_empty else 0
        score += text_ratio * 0.30

        # Signal 2: Uniqueness of values
        unique_count = len(set(non_empty))
        uniqueness = unique_count / len(non_empty) if non_empty else 0
        score += uniqueness * 0.25

        # Signal 3: Fill rate (non-empty in candidate row)
        fill_rate = len(non_empty) / region.column_count if region.column_count > 0 else 0
        score += fill_rate * 0.15

        # Signal 4: Data type consistency below
        if region.row_count > 1:
            consistency_score = self._check_type_consistency_below(
                sheet, region, candidate_row
            )
            score += consistency_score * 0.20

        # Signal 5: Shorter cell values than data below (headers tend to be labels)
        if non_empty and region.row_count > 1:
            header_avg_len = sum(len(v) for v in non_empty) / len(non_empty)
            data_avg_len = self._avg_data_length_below(sheet, region, candidate_row)
            if data_avg_len > 0 and header_avg_len <= data_avg_len * 1.5:
                score += 0.10
            elif header_avg_len <= 50:
                score += 0.05

        return min(1.0, score)

    def _check_type_consistency_below(
        self, sheet: Any, region: DetectedRegion, header_row: int
    ) -> float:
        """Check if data rows below the candidate have consistent types per column.

        Args:
            sheet: The worksheet.
            region: The region.
            header_row: 0-based candidate header row.

        Returns:
            Score 0.0 to 1.0 indicating consistency.
        """
        data_start = header_row + 1
        rows_to_check = min(10, region.end_row - data_start)

        if rows_to_check <= 0:
            return 0.0

        # For each column, check type consistency
        consistent_cols = 0

        for col_offset in range(region.column_count):
            col_1based = region.start_column + col_offset + 1
            types_seen: set[str] = set()

            for row_offset in range(rows_to_check):
                row_1based = data_start + row_offset + 1
                cell = sheet.cell(row=row_1based, column=col_1based)
                value = self._get_cell_value(cell)
                if value is not None:
                    types_seen.add(self._classify_value_type(value))

            # A consistent column has 1-2 types (e.g., numeric + None)
            if len(types_seen) <= 2:
                consistent_cols += 1

        return consistent_cols / region.column_count if region.column_count > 0 else 0

    def _avg_data_length_below(
        self, sheet: Any, region: DetectedRegion, header_row: int
    ) -> float:
        """Calculate average cell content length in data rows below the header.

        Args:
            sheet: The worksheet.
            region: The region.
            header_row: 0-based candidate header row.

        Returns:
            Average string length of data cells.
        """
        data_start = header_row + 1
        rows_to_check = min(10, region.end_row - data_start)

        if rows_to_check <= 0:
            return 0.0

        total_length = 0
        cell_count = 0

        for row_offset in range(rows_to_check):
            row_1based = data_start + row_offset + 1
            for col_offset in range(region.column_count):
                col_1based = region.start_column + col_offset + 1
                cell = sheet.cell(row=row_1based, column=col_1based)
                value = self._get_cell_value(cell)
                if value is not None:
                    total_length += len(str(value))
                    cell_count += 1

        return total_length / cell_count if cell_count > 0 else 0.0

    def _explain_header_selection(
        self, candidate_values: list[str], score: float, offset: int
    ) -> str:
        """Generate a human-readable explanation for header selection.

        Args:
            candidate_values: Values in the selected header row.
            score: The detection score.
            offset: Row offset from region start.

        Returns:
            Explanation string.
        """
        non_empty = [v for v in candidate_values if v.strip()]
        if offset == 0:
            return (
                f"First row selected as header (score={score:.2f}): "
                f"{len(non_empty)} non-empty cells with text content"
            )
        return (
            f"Row at offset {offset} selected as header (score={score:.2f}): "
            f"better text density and uniqueness than preceding rows"
        )

    def _infer_schema(
        self, sheet: Any, region: DetectedRegion, header_row: int
    ) -> list[ColumnSchema]:
        """Infer column types from content below the header.

        Examines data rows to determine column types: string, integer, decimal,
        boolean, date, datetime, or unknown. Does not aggressively convert
        ambiguous values.

        Args:
            sheet: The worksheet.
            region: The region to infer schema for.
            header_row: 0-based row index of the header.

        Returns:
            List of ColumnSchema for each column in the region.
        """
        columns: list[ColumnSchema] = []
        data_start = header_row + 1
        rows_to_check = min(50, region.end_row - data_start)

        for col_offset in range(region.column_count):
            col_1based = region.start_column + col_offset + 1
            header_1based = header_row + 1

            # Get column name from header
            header_cell = sheet.cell(row=header_1based, column=col_1based)
            header_value = self._get_cell_value(header_cell)
            col_name = (
                str(header_value).strip()
                if header_value is not None
                else f"Column_{col_offset + 1}"
            )
            if not col_name:
                col_name = f"Column_{col_offset + 1}"

            # Collect values and types from data rows
            type_counts: dict[str, int] = {}
            sample_values: list[str] = []
            has_null = False
            total_values = 0

            for row_offset in range(rows_to_check):
                row_1based = data_start + row_offset + 1
                if row_1based > (region.end_row):
                    break
                cell = sheet.cell(row=row_1based, column=col_1based)
                value = self._get_cell_value(cell)

                if value is None:
                    has_null = True
                    continue

                total_values += 1
                value_type = self._classify_value_type(value)
                type_counts[value_type] = type_counts.get(value_type, 0) + 1

                # Collect up to 5 sample values
                if len(sample_values) < 5:
                    sample_values.append(str(value))

            # Determine dominant type
            data_type, confidence = self._resolve_column_type(type_counts, total_values)

            columns.append(
                ColumnSchema(
                    name=col_name,
                    data_type=data_type,
                    nullable=has_null,
                    column_index=col_offset,
                    sample_values=sample_values,
                    confidence=confidence,
                )
            )

        return columns

    def _resolve_column_type(
        self, type_counts: dict[str, int], total_values: int
    ) -> tuple[str, float]:
        """Determine the dominant data type for a column.

        Args:
            type_counts: Count of each detected type.
            total_values: Total non-null values examined.

        Returns:
            Tuple of (data_type, confidence).
        """
        if total_values == 0:
            return "unknown", 0.0

        # Find the most common type
        dominant_type = max(type_counts, key=type_counts.get)  # type: ignore[arg-type]
        dominant_count = type_counts[dominant_type]
        confidence = dominant_count / total_values

        # If confidence is too low, mark as string (safe fallback)
        if confidence < 0.6:
            return "string", confidence

        return dominant_type, round(confidence, 4)

    def _classify_value_type(self, value: Any) -> str:
        """Classify a single cell value into a type category.

        Args:
            value: The cell value to classify.

        Returns:
            One of: string, integer, decimal, boolean, date, datetime, unknown.
        """
        if value is None:
            return "unknown"

        if isinstance(value, bool):
            return "boolean"

        if isinstance(value, int):
            return "integer"

        if isinstance(value, float):
            # Check if it's actually an integer value stored as float
            if value == int(value) and not (value > 2**53 or value < -(2**53)):
                return "integer"
            return "decimal"

        if isinstance(value, datetime):
            return "datetime"

        if isinstance(value, date):
            return "date"

        if isinstance(value, time):
            return "datetime"

        # String value — check if it looks like a specific type
        str_val = str(value).strip()
        if not str_val:
            return "string"

        # Check boolean-like strings
        if str_val.lower() in ("true", "false", "yes", "no", "y", "n"):
            return "boolean"

        # Check numeric strings
        cleaned = str_val.replace(",", "").replace("$", "").replace("%", "")
        try:
            num = float(cleaned)
            if "." not in cleaned and num == int(num):
                return "integer"
            return "decimal"
        except (ValueError, OverflowError):
            pass

        return "string"

    def _extract_content_sample(
        self,
        sheet: Any,
        start_row_1based: int,
        end_row_1based: int,
        start_col_1based: int,
        end_col_1based: int,
    ) -> list[list[str]]:
        """Extract up to _MAX_SAMPLE_ROWS rows of content as string lists.

        Args:
            sheet: The worksheet.
            start_row_1based: First row (1-based, inclusive).
            end_row_1based: Last row (1-based, inclusive).
            start_col_1based: First column (1-based, inclusive).
            end_col_1based: Last column (1-based, inclusive).

        Returns:
            List of row lists with string cell values.
        """
        sample: list[list[str]] = []
        rows_to_sample = min(_MAX_SAMPLE_ROWS, end_row_1based - start_row_1based + 1)

        for row_offset in range(rows_to_sample):
            row_idx = start_row_1based + row_offset
            row_values: list[str] = []
            for col_idx in range(start_col_1based, end_col_1based + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = self._get_cell_value(cell)
                row_values.append(str(value) if value is not None else "")
            sample.append(row_values)

        return sample

    def _build_raw_text(
        self,
        sheet: Any,
        start_row_1based: int,
        end_row_1based: int,
        start_col_1based: int,
        end_col_1based: int,
    ) -> str:
        """Build raw text from region cells for narrative classification.

        Args:
            sheet: The worksheet.
            start_row_1based: First row (1-based, inclusive).
            end_row_1based: Last row (1-based, inclusive).
            start_col_1based: First column (1-based, inclusive).
            end_col_1based: Last column (1-based, inclusive).

        Returns:
            Concatenated text content with newlines between rows.
        """
        lines: list[str] = []

        for row_idx in range(start_row_1based, end_row_1based + 1):
            row_parts: list[str] = []
            for col_idx in range(start_col_1based, end_col_1based + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = self._get_cell_value(cell)
                if value is not None:
                    text = str(value).strip()
                    if text:
                        row_parts.append(text)
            if row_parts:
                lines.append(" ".join(row_parts))

        return "\n".join(lines)

    @staticmethod
    def _get_cell_value(cell: Any) -> Any:
        """Get the resolved value of a cell, handling merged cells and formulas.

        For read_only + data_only mode, formulas return their cached value.
        MergedCell instances return None (only the top-left cell has the value).

        Args:
            cell: An openpyxl Cell or MergedCell.

        Returns:
            The cell's value, or None for empty/merged cells.
        """
        if isinstance(cell, MergedCell):
            return None

        value = cell.value
        if value is None:
            return None

        # Strip whitespace from string values
        if isinstance(value, str):
            return value if value.strip() else None

        return value

    @staticmethod
    def _looks_numeric(value: str) -> bool:
        """Check if a string value looks like a number.

        Args:
            value: String to check.

        Returns:
            True if the value appears numeric.
        """
        cleaned = value.strip().replace(",", "").replace("$", "").replace("%", "")
        if not cleaned:
            return False
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
